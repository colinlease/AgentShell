from __future__ import annotations

import csv
import io
import json
import re
import zipfile
from pathlib import Path
from typing import Any
from xml.etree import ElementTree

from openpyxl import load_workbook

from app.workspace_apps.local_knowledge.services.index_store import LocalKnowledgeIndexStore


MAX_EXTRACTABLE_FILE_BYTES = 25 * 1024 * 1024
DEFAULT_MAX_CHARS = 6000
MAX_CHARS_LIMIT = 20000
MAX_SEARCH_TEXT_CHARS = 200000


class LocalKnowledgeContentError(RuntimeError):
    """Raised when a file excerpt cannot be safely extracted."""


def read_file_excerpt(
    *,
    root_path: str,
    relative_path: str,
    max_chars: int | None = DEFAULT_MAX_CHARS,
) -> dict[str, Any]:
    normalized_path = _normalize_relative_path(relative_path)
    if not normalized_path:
        raise LocalKnowledgeContentError("path is required.")

    store = LocalKnowledgeIndexStore(root_path=str(root_path))
    record = store.get_file(normalized_path)
    if record is None:
        raise LocalKnowledgeContentError("File is not present in the current Local Knowledge inventory.")
    if str(record.get("support_status")) != "supported":
        raise LocalKnowledgeContentError("File type is visible in inventory but is not supported for content excerpts.")

    file_path = _resolve_inventory_path(root_path=str(root_path), relative_path=normalized_path)
    size_bytes = int(record.get("size_bytes") or 0)
    if size_bytes > MAX_EXTRACTABLE_FILE_BYTES:
        raise LocalKnowledgeContentError(
            f"File is too large for bounded excerpt extraction ({size_bytes} bytes)."
        )

    limit = _normalize_max_chars(max_chars)
    extension = str(record.get("extension") or file_path.suffix.lower())
    extraction = _extract_by_extension(file_path=file_path, extension=extension, max_chars=limit)
    content = _truncate_text(str(extraction["content"]), limit)
    return {
        "status": "ok",
        "path": normalized_path,
        "name": str(record.get("name") or file_path.name),
        "extension": extension,
        "kind": str(record.get("kind") or ""),
        "size_bytes": size_bytes,
        "mtime_ns": int(record.get("mtime_ns") or 0),
        "extraction_method": extraction["method"],
        "content": content["text"],
        "truncated": bool(extraction.get("truncated", False) or content["truncated"]),
        "read_only": True,
    }


def extract_searchable_text(
    *,
    root_path: str,
    relative_path: str,
    max_chars: int | None = MAX_SEARCH_TEXT_CHARS,
) -> dict[str, Any]:
    normalized_path = _normalize_relative_path(relative_path)
    if not normalized_path:
        raise LocalKnowledgeContentError("path is required.")

    store = LocalKnowledgeIndexStore(root_path=str(root_path))
    record = store.get_file(normalized_path)
    if record is None:
        raise LocalKnowledgeContentError("File is not present in the current Local Knowledge inventory.")
    if str(record.get("support_status")) != "supported":
        raise LocalKnowledgeContentError("File type is visible in inventory but is not supported for content indexing.")

    file_path = _resolve_inventory_path(root_path=str(root_path), relative_path=normalized_path)
    size_bytes = int(record.get("size_bytes") or 0)
    if size_bytes > MAX_EXTRACTABLE_FILE_BYTES:
        raise LocalKnowledgeContentError(
            f"File is too large for content indexing ({size_bytes} bytes)."
        )

    limit = _normalize_search_max_chars(max_chars)
    extension = str(record.get("extension") or file_path.suffix.lower())
    extraction = _extract_by_extension(file_path=file_path, extension=extension, max_chars=limit)
    content = _truncate_text(str(extraction["content"]), limit)
    return {
        "status": "ok",
        "path": normalized_path,
        "content_hash": str(record.get("content_hash") or ""),
        "extension": extension,
        "extraction_method": extraction["method"],
        "content": content["text"],
        "truncated": bool(extraction.get("truncated", False) or content["truncated"]),
    }


def _extract_by_extension(*, file_path: Path, extension: str, max_chars: int) -> dict[str, Any]:
    try:
        if extension in {".txt", ".md", ".py", ".json"}:
            return {
                "method": "text",
                "content": _read_text_file(file_path, max_chars=max_chars),
                "truncated": False,
            }
        if extension == ".csv":
            return {
                "method": "csv_preview",
                "content": _read_csv_preview(file_path, max_chars=max_chars),
                "truncated": False,
            }
        if extension in {".xlsx", ".xlsm"}:
            return {
                "method": "excel_workbook_preview",
                "content": _read_excel_preview(file_path),
                "truncated": False,
            }
        if extension == ".docx":
            return {
                "method": "docx_text",
                "content": _read_docx_text(file_path),
                "truncated": False,
            }
        if extension == ".pptx":
            return {
                "method": "pptx_text",
                "content": _read_pptx_text(file_path),
                "truncated": False,
            }
        if extension == ".pdf":
            return {
                "method": "pdf_text_optional",
                "content": _read_pdf_text(file_path),
                "truncated": False,
            }
        if extension == ".xls":
            raise LocalKnowledgeContentError("Legacy .xls excerpts are not available without an additional parser.")
        raise LocalKnowledgeContentError(f"Unsupported file extension for content excerpts: {extension or 'none'}.")
    except LocalKnowledgeContentError:
        raise
    except zipfile.BadZipFile as exc:
        raise LocalKnowledgeContentError(
            f"Could not extract content from this {extension or 'file'} file because it is not a valid archive."
        ) from exc
    except Exception as exc:
        raise LocalKnowledgeContentError(
            f"Could not extract content from this {extension or 'file'} file."
        ) from exc


def _resolve_inventory_path(*, root_path: str, relative_path: str) -> Path:
    root = Path(root_path).resolve()
    candidate = (root / relative_path).resolve()
    try:
        candidate.relative_to(root)
    except ValueError as exc:
        raise LocalKnowledgeContentError("path must stay inside the mounted folder.") from exc
    if not candidate.exists() or not candidate.is_file():
        raise LocalKnowledgeContentError("File is not available on disk.")
    return candidate


def _read_text_file(file_path: Path, *, max_chars: int) -> str:
    byte_limit = min(max(max_chars * 4, 65536), MAX_EXTRACTABLE_FILE_BYTES)
    raw = file_path.read_bytes()[:byte_limit]
    return _decode_text(raw)


def _read_csv_preview(file_path: Path, *, max_chars: int) -> str:
    text = _read_text_file(file_path, max_chars=max_chars)
    rows: list[list[str]] = []
    reader = csv.reader(io.StringIO(text))
    for index, row in enumerate(reader):
        if index >= 30:
            break
        rows.append([str(value) for value in row[:20]])
    if not rows:
        return ""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return output.getvalue().strip()


def _read_excel_preview(file_path: Path) -> str:
    workbook = load_workbook(file_path, read_only=True, data_only=True)
    sections: list[str] = []
    try:
        for sheet_name in workbook.sheetnames[:8]:
            sheet = workbook[sheet_name]
            sections.append(f"Sheet: {sheet_name}")
            sections.append(f"Dimensions: rows={sheet.max_row}, columns={sheet.max_column}")
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 25), max_col=min(sheet.max_column, 12)):
                values = [_format_cell_value(cell.value) for cell in row]
                if any(value for value in values):
                    sections.append(" | ".join(values))
            sections.append("")
    finally:
        workbook.close()
    return "\n".join(sections).strip()


def _read_docx_text(file_path: Path) -> str:
    with zipfile.ZipFile(file_path) as archive:
        names = ["word/document.xml"]
        names.extend(name for name in archive.namelist() if name.startswith("word/header") or name.startswith("word/footer"))
        chunks: list[str] = []
        for name in names:
            if name not in archive.namelist():
                continue
            chunks.extend(_text_from_xml_bytes(archive.read(name), text_tag_suffix="}t"))
    return _normalize_extracted_text("\n".join(chunks))


def _read_pptx_text(file_path: Path) -> str:
    with zipfile.ZipFile(file_path) as archive:
        slide_names = sorted(
            name
            for name in archive.namelist()
            if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        sections: list[str] = []
        for index, name in enumerate(slide_names, start=1):
            text = _normalize_extracted_text("\n".join(_text_from_xml_bytes(archive.read(name), text_tag_suffix="}t")))
            if text:
                sections.append(f"Slide {index}:\n{text}")
    return "\n\n".join(sections).strip()


def _read_pdf_text(file_path: Path) -> str:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as exc:
        raise LocalKnowledgeContentError(
            "PDF text extraction is not available in this environment yet."
        ) from exc

    reader = PdfReader(str(file_path))
    pages: list[str] = []
    for page in reader.pages[:20]:
        page_text = page.extract_text() or ""
        if page_text:
            pages.append(page_text)
    return _normalize_extracted_text("\n\n".join(pages))


def _text_from_xml_bytes(xml_bytes: bytes, *, text_tag_suffix: str) -> list[str]:
    try:
        root = ElementTree.fromstring(xml_bytes)
    except ElementTree.ParseError:
        return []
    values: list[str] = []
    for element in root.iter():
        if element.tag.endswith(text_tag_suffix) and element.text:
            values.append(element.text)
    return values


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _normalize_extracted_text(text: str) -> str:
    lines = [re.sub(r"\s+", " ", line).strip() for line in str(text or "").splitlines()]
    return "\n".join(line for line in lines if line).strip()


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=True)
    return str(value).replace("\n", " ").strip()


def _truncate_text(text: str, max_chars: int) -> dict[str, Any]:
    if len(text) <= max_chars:
        return {"text": text, "truncated": False}
    return {"text": text[:max_chars].rstrip(), "truncated": True}


def _normalize_max_chars(max_chars: int | None) -> int:
    if max_chars is None:
        return DEFAULT_MAX_CHARS
    try:
        value = int(max_chars)
    except (TypeError, ValueError):
        return DEFAULT_MAX_CHARS
    return max(500, min(value, MAX_CHARS_LIMIT))


def _normalize_search_max_chars(max_chars: int | None) -> int:
    if max_chars is None:
        return MAX_SEARCH_TEXT_CHARS
    try:
        value = int(max_chars)
    except (TypeError, ValueError):
        return MAX_SEARCH_TEXT_CHARS
    return max(1000, min(value, MAX_SEARCH_TEXT_CHARS))


def _normalize_relative_path(path: str | None) -> str:
    value = str(path or "").strip().replace("\\", "/").strip("/")
    if value in {"", "."}:
        return ""
    parts = [part for part in value.split("/") if part]
    if any(part in {".", ".."} for part in parts):
        raise LocalKnowledgeContentError("path must be a folder-relative file path without '.' or '..' segments.")
    if parts and ":" in parts[0]:
        raise LocalKnowledgeContentError("path must be relative to the mounted folder, not an absolute path.")
    return "/".join(parts)
